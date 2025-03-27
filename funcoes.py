from databases import BDS, Bawm, PosicaoDm1Pickle, Crm
from funcoes_datas import FuncoesDatas

import xlwings as xw
from datetime import date,timedelta
from dateutil.relativedelta import relativedelta
import pandas as pd
import math
import copy
import re
import numpy as np

from scipy.stats import norm

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import range_boundaries

bds = BDS()
bawm = Bawm()


def ewma_zscore(serie:pd.Series, alpha=0.1, janela:int=12) -> pd.Series:
    """
    Calcula o z-score de uma série fazendo a média móvel em EWMA

    Parameters
    ----------
    serie : pd.Series
        Serie do pandas com a série temporal
    alpha : TYPE, optional
        Parâmetro para o EWMA. The default is 0.1.
    janela : int, optional
        Número de períodos para calcular a variância. The default is 12.

    Returns
    -------
    retorno : pd.Series
        Mesma série, ajustada.

    """
    variancia = serie.rolling(janela).var()
    media = serie.ewm(alpha=alpha, adjust=False).mean()
    retorno = pd.Series(index=serie.index, name=serie.name)
    var_ewma = pd.Series(index=serie.index)
    var_ant = None
    for idx in serie.index:
        if not np.isnan(variancia[idx]) and not np.isnan(media[idx]):
            if not var_ant:
                var_ant = variancia[idx]
                var_ewma[idx] = var_ant
                retorno[idx] = None
            else:
                var_ewma[idx] = (1-alpha) * (var_ant + alpha * (serie[idx] - media[idx]) ** 2)
                var_ant = var_ewma[idx]
                retorno[idx] = (serie[idx] - media[idx]) / (var_ewma[idx] ** 0.5)
    return retorno

def e_numero(valor):
    """
    Determina se valor é numérico ou não

    Parameters
    ----------
    valor : string ou numerico (ou qualquer outra coisa)
        DESCRIPTION.

    Returns
    -------
    bool
        True se for númerico.

    """
    try:
        if not valor:
            return False
        if re.match(r'^[-+]?[0-9]+\.?[0-9]*$', valor):
            return True
        elif re.match(r'^[-+]?[0-9]+\,?[0-9]*$', valor):
            return True
        else:
            return False
    except:
        return False
    

#support function: read excel ranges 
def get_range(path, label):
    wb = load_workbook(path, read_only= True, data_only= True)     
    ws, wrange = next(wb.defined_names[label].destinations)
    materials = [[cell.value for cell in row] for row in wb[ws][wrange]]
    
    out = pd.DataFrame(materials[1:], columns= materials[0])
    out = out.dropna(how='all', axis= 0)
    
    wb.close()
    return out

#function to fill named range with dataframe by starting position
def fill_range_df(path, label, table, keep_index=False, keep_header=True, sheet_name:str=None):
    wb = load_workbook(path, read_only= False)
    ws = wb.active
    if sheet_name:
        ws = wb[sheet_name]
    
    area = wb.defined_names[label]
    rng_string = area.attr_text.split('!')[1]
    rng_bounds = range_boundaries(rng_string) 

    stt_row = rng_bounds[1] - (1 if keep_header else 2)
    stt_col = rng_bounds[0] - 1

    rows = dataframe_to_rows(table, index= keep_index, header= keep_header)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row= r_idx +stt_row, 
                    column= c_idx +stt_col,
                    value= value)
            
    wb.save(path)
    wb.close()

def fill_range_value(path_copy, label, value):
    wb = load_workbook(path_copy, read_only= False)
    ws = wb.active
    
    area = wb.defined_names[label]
    rng_string = area.attr_text.split('!')[1]
    rng_bounds = range_boundaries(rng_string) 

    stt_row = rng_bounds[1]
    stt_col = rng_bounds[0]

    ws.cell(row= stt_row, column= stt_col, value= value)
    
    wb.save(path_copy)
    wb.close()


def pd_reorder_columns(df:pd.DataFrame, preferred_columns:list) -> pd.DataFrame:
    """
    Reorders the columns of a DataFrame, placing the specified columns at the beginning.
    
    Parameters:
    - df (pd.DataFrame): Input DataFrame.
    - preferred_columns (list): List of column names to place at the beginning.
    
    Returns:
    - pd.DataFrame: DataFrame with reordered columns.
    """
    # Ensure all preferred columns exist in the DataFrame
    if not set(preferred_columns).issubset(set(df.columns)):
        raise ValueError("All preferred columns must exist in the DataFrame.")
    
    # Split the columns into two lists: preferred and non-preferred
    other_columns = [col for col in df.columns if col not in preferred_columns]
    
    # Combine the preferred and non-preferred columns, maintaining their original order within each group
    new_column_order = preferred_columns + other_columns
    
    # Return the DataFrame with the reordered columns
    return df[new_column_order]

def drawdown(serie:pd.Series, base_cota:bool=True) -> pd.Series:
    """
    Calcula o drawdown de uma série de cotas ou retornos

    Parameters
    ----------
    serie : pd.Series
        Serie a ser utilizada para o calculo.
    base_cota : bool, optional
        Indica se serie é uma cota ou retornos. The default is True (Cota).

    Returns
    -------
    retorno : pd.Series
        Serie de drawdowns.

    """
    if not base_cota:
        cota = serie.copy().fillna(0)
        cota = cota.apply(lambda x: x+1)
        cota = cota.cumprod()
    else:
        cota = serie.copy()
    topo = cota.cummax()
    retorno = cota / topo - 1
    return retorno


def nelson_siegel_yield_curve(maturities:np.array, betas:np.array, lambdas:np.array, maturities_em_anos:bool=True):
    """
    Implementa a função de Nelson Siegel de acordo com a metodologia utilizada nas curvas interpoladas
    da ANBIMA

    Parameters
    ----------
    maturities : np.array
        Prazos em dias úteis (dentro da função matriz é dividida por 252).
    betas : np.array
        DESCRIPTION.
    lambdas : np.array
        DESCRIPTION.
    maturities_em_anos: bool
        Se False divide prazos por 252
        Usar False se input estiver em dias úteis
    Returns
    -------
    yield_curve : np.array
        Curva interpolada.

    """
    beta1, beta2, beta3, beta4 = betas
    lambda1, lambda2 = lambdas
    if maturities_em_anos:
        mat_adj = maturities
    else:
        mat_adj = maturities / 252
    term1 = beta1
    term2 = beta2 * (1 - np.exp(-mat_adj * lambda1)) / (mat_adj * lambda1)
    term3 = beta3 * ((1 - np.exp(-mat_adj * lambda1)) / (mat_adj * lambda1) - np.exp(-mat_adj * lambda1))
    term4 = beta4 * ((1 - np.exp(-mat_adj * lambda2)) / (mat_adj * lambda2) - np.exp(-mat_adj * lambda2))
    yield_curve = term1 + term2 + term3 + term4
    return yield_curve

    
class CalculosRiscoPort:
    """
    Classe criada para fazer estimativa da pontuação de risco da supercarteira via 
    volatilidade das classes de ativos
    """
    def __init__(self, carregar=True, homologacao=False, offshore = False):        
        
        self.bawm = Bawm(homologacao=homologacao)
        
        if offshore:
            self.dicio_class_ind = {'alternativos crédito': 'HFRI Cred ',
                                         'alternativos macro arbitragem': 'HFRI Macro',
                                         'alternativos long short': 'HFRI RV   ',
                                         'alternativos macro arbitragem': 'HFRI Macro',
                                         'iliq. p. equity': 'Russel2000',
                                         'real estate': 'Ftse_Reits',
                                         'outros': 'NDUEACWF  ',
                                         'liquidez': 'CDI',
                                         'renda fixa global': 'BBG_BARC  ',
                                         'renda fixa crédito': 'RF_LIVRE',
                                         'renda variável global': 'NDUEACWF  '}

            dicio = [{'Pontos': 250, 'Stress': 0.043},
                     {'Pontos': 400, 'Stress': 0.069},
                     {'Pontos': 600, 'Stress': 0.104},
                     {'Pontos': 800, 'Stress': 0.139},
                     {'Pontos': 1000, 'Stress': 0.173},
                     {'Pontos': 1500, 'Stress': 0.26},
                     {'Pontos': 2000, 'Stress': 0.346}]
        else:    
            self.dicio_class_ind = {'R Fixa Pós': 'CDI','R Fixa Pré': 'IRFM','R Fixa Infl': 'IMAIPCA',
                                    'RF Internacional': 'BMLUSHY','Alternativos': 'IHFA','R Variável': 'Ibvsp',
                                    'Real Estate': 'IFIX','RV Internacional': 'NDUEACWF;PTAXV','Crédito Alternativo': 'CDI',
                                    'P. Equity': 'SMLLBVA','Outros': 'Ibvsp','Cambial': 'PTAXV',
                                    'Infraestrutura Ilíquidos': 'SMLLBVA','RF Distressed': 'SMLLBVA','F Exclusivo': 'IHFA',
                                    'Investimento no Exterior': 'NDUEACWF;PTAXV','RF Infl': 'IHFA','Imobiliário': 'IFIX',
                                    'F Exclusivo Ações': 'Ibvsp'}
            dicio = [{'Pontos': 250, 'Stress': 0.043},
                     {'Pontos': 400, 'Stress': 0.069},
                     {'Pontos': 600, 'Stress': 0.104},
                     {'Pontos': 800, 'Stress': 0.139},
                     {'Pontos': 1000, 'Stress': 0.173},
                     {'Pontos': 1500, 'Stress': 0.26},
                     {'Pontos': 2000, 'Stress': 0.346}]
        self.regua = pd.DataFrame(dicio)
        self.sigma = 7.33 

        self.cotas = pd.DataFrame()
        self.correl = pd.DataFrame()
        self.vols = pd.DataFrame()
        
        if carregar:
            self.busca_cotas()
            self.vols, self.correl = self.vol_correl()
    
    
    def busca_cotas(self):
        data_fim = self.bawm.banco.hoje() - relativedelta(days=5)
        data_ini = data_fim - relativedelta(days=1500)
        
        cotas = []
        for item in self.dicio_class_ind.items():
            classe = item[0]
            indice = item[1]
            # Separação entre índices compostos
            if ';' in indice:
                indice = indice.split(';')
            else:
                indice = [indice]
            if len(indice) == 1:
                df = self.bawm.in_historico(index_cod=indice[0], data_ini=data_ini, data_fim=data_fim).drop('DataAtu', axis=1)
            else:
                df = []
                for ind in indice:
                    df.append(self.bawm.in_historico(index_cod=ind, data_ini=data_ini, data_fim=data_fim).drop('DataAtu', axis=1))
                df = pd.concat(df)
                df = pd.pivot_table(df, values='Valor', index='Data', columns='IndexCod').fillna(method='ffill')
                df['Valor'] = df[indice[0]] * df[indice[1]]
                df.reset_index(inplace=True)
            df.insert(0, 'Classe', [classe] * len(df))
            cotas.append(df)
        cotas = pd.concat(cotas)
        cotas = pd.pivot_table(cotas, values='Valor', index='Data', columns='Classe').fillna(method='ffill')
        cotas.replace(0, method='ffill', inplace=True)
        self.cotas = cotas
    
    def vol_correl(self, dias=504):
        data_fim = self.bawm.banco.hoje() - relativedelta(days=5)
        data_ini = data_fim - relativedelta(days=370)
               
        retornos = self.cotas / self.cotas.shift(1) - 1
        retornos.dropna(inplace=True)
        retornos = retornos.iloc[-dias:]
        vols = retornos.std() * (252 ** 0.5)
        vols['R Fixa Pós'] = 0.02 # Aumenta vol da renda fixa pós
        correl = retornos.corr()

        #forçando rv global hedged
        if 'RV Global Hedged' not in vols.index:
            vols['RV Global Hedged'] = vols['RV Internacional']
            correl.loc['RV Global Hedged'] = correl.loc['RV Internacional'].copy()
            correl['RV Global Hedged'] = correl['RV Internacional']
        return vols, correl
    
    def calcula_vols(self, df_aloc):
        campos = ['PesoFinal', 'PesoFuturo', 'PesoProjetado']
        resultado = []
        for campo in campos:
            if campo in df_aloc.columns:
                dicio = {}        
                vol = self.__calcula_vol__(df_aloc, campo)                
                pts = self.interpola_pontos_risco(vol)
                te = None
                if 'Atual' in df_aloc.columns:
                    df_temp = df_aloc.copy().fillna(0)
                    df_temp.insert(len(df_temp.columns), 'Diff', [0] * len(df_temp))
                    df_temp['Diff'] = df_temp.apply(lambda x: x['Atual'] - x[campo], axis=1)
                    te = self.__calcula_vol__(df_temp, 'Diff') 
                dicio = {'Campo': campo, 'Vol': vol, 'Pontos': pts, 'TE': te}
                resultado.append(dicio)
        return resultado
    
    def __calcula_vol__(self, df_aloc, campo):        
        teste = df_aloc.copy().reset_index()
        classes = list(teste['Classe'].to_list())
        vols = self.vols[classes]
        correl = self.correl.loc[classes][classes]
        teste.set_index('Classe', inplace=True)
                
        teste['Vol'] = vols
        teste['CTR'] = teste[campo] * teste['Vol']            
        teste = teste[['CTR']]
        teste = pd.concat([teste, correl], axis=1).dropna()
    
        # Calculo
        ctr = teste['CTR'].to_numpy()
        correl = teste.drop('CTR', axis=1).to_numpy()
        vol_port = np.matmul(np.transpose(ctr),(np.matmul(correl,ctr)))
        return vol_port ** 0.5

    def interpola_pontos_risco(self, vol):
        from sklearn.linear_model import LinearRegression
        y = self.regua['Pontos'].to_numpy()
        x = self.regua['Stress'].to_numpy()
        x = x.reshape((-1, 1))
        stress = vol / (12 ** 0.5) * self.sigma
        model = LinearRegression().fit(x, y)
        return float(stress * model.coef_ + model.intercept_)
    
    def vol_carteira(self,df):
            
            classes = list(df['Classe'].to_list())
            vols = self.vols[classes]
            correl = self.correl.loc[classes][classes]
            df.set_index('Classe', inplace=True)
            df['Vol'] = vols
            df['CTR'] = df['Atual'] * df['Vol']
            delta = df['Atual'].to_numpy()
            df = df[['CTR']]
            df = pd.concat([df, correl], axis=1).dropna()
            
            # Calculo
            ctr = df['CTR'].to_numpy()
            correl = df.drop('CTR', axis=1).to_numpy()
            te = np.dot(np.dot(ctr,correl),ctr)
            vol_port = np.matmul(np.transpose(ctr),(np.matmul(correl,ctr)))
            return vol_port ** 0.5    
    

    def tracking_error(self,df):
            
            #df = self.asset_class_allocation()
            
            #teste = df.copy()
            classes = list(df['Classe'].to_list())
            vols = self.vols[classes]
            correl = self.correl.loc[classes][classes]
            df.set_index('Classe', inplace=True)
            df['Delta'] = df['Tático'] - df['Atual']
            df['Vol'] = vols
            df['CTR'] = df['Delta'] *df['Vol']
            delta = df['Delta'].to_numpy()
            df = df[['CTR']]
            df = pd.concat([df, correl], axis=1).dropna()
            
            # Calculo
            ctr = df['CTR'].to_numpy()
            correl = df.drop('CTR', axis=1).to_numpy()
            te = np.dot(np.dot(ctr,correl),ctr)
            vol_port = np.matmul(np.transpose(ctr),(np.matmul(correl,ctr)))
            return vol_port ** 0.5


class Option:
    
    def __init__(self, type, strike, dividend, volatility,
                 stock_price, interest_rate, time_to_maturity):
        self.type = type
        self.strike = strike
        self.dividend = dividend
        self.volatility = volatility
        self.stock_price = stock_price
        self.interest_rate = interest_rate
        self.time_to_maturity = time_to_maturity

    ############################################################################
    def __repr__(self):
        return ("Option: {"
            + "\n  type: " + str(self.type)
            + "\n  strike: " + str(self.strike)
            + "\n  dividend: " + str(self.dividend)
            + "\n  volatility: " + str(self.volatility)
            + "\n  stock_price: " + str(self.stock_price)
            + "\n  interest_rate: " + str(self.interest_rate)
            + "\n  time_to_maturity: " + str(self.time_to_maturity)
            + "\n  d1: " + str(self.d1)
            + "\n  d2: " + str(self.d2)
            + "\n  delta: " + str(self.delta)
            + "\n  gamma: " + str(self.gamma)
            + "\n  vega: " + str(self.vega)
            + "\n  theta: " + str(self.theta)
            + "\n  price: " + str(self.price) + "\n}"
        )

    ############################################################################
    def reset(self):
        self._d1 = None
        self._d2 = None
        self._price = None
        self._delta = None
        self._gamma = None
        self._vega = None
        self._theta = None
        self._parity = None

    ############################################################################
    # Plot                                                                     #
    ############################################################################
    @staticmethod
    def plot(options, x_axis, y_axis, x_start, x_end, x_samples, save_to):
        is_inverted = False
        if (x_end < x_start):
            x_start, x_end = x_end, x_start
            is_inverted = True

        x_values = np.linspace(x_start, x_end, x_samples)
        y_values = {}

        for (i, option) in enumerate(options):
            option = copy.copy(option)
            key = str(i+1) + ". " + option.type
            y_values[key] = []
            for x in x_values:
                setattr(option, x_axis, x)
                y = getattr(option, y_axis)
                y_values[key].append(y)

        df = pd.DataFrame(data=y_values, index=x_values)

        fig = df.plot(style=[
            ("-" if (option.type == "CALL") else "--") for option in options
        ])
        if (is_inverted): fig.invert_xaxis()
        # plt.title("Options: " + x_axis + " x " + y_axis)
        # plt.xlabel(x_axis)
        # plt.ylabel(y_axis)
        # plt.grid()
        # plt.legend(loc=(1.05, 0.5))
        # plt.tight_layout()
        # plt.savefig(save_to)

    ############################################################################
    # Inputs                                                                   #
    ############################################################################
    @property
    def type(self):
        return self._type

    @type.setter
    def type(self, value):
        if (value != "CALL" and value != "PUT"):
            raise ValueError("Invalid option type!")
        self._type = value
        self.reset()

    ############################################################################
    @property
    def strike(self):
        return self._strike

    @strike.setter
    def strike(self, value):
        if (value <= 0):
            raise ValueError("Invalid strike!")
        self._strike = value
        self.reset()

    ############################################################################
    @property
    def dividend(self):
        return self._dividend

    @dividend.setter
    def dividend(self, value):
        if (value < 0):
            raise ValueError("Invalid dividend!")
        self._dividend = value
        self.reset()

    ############################################################################
    @property
    def volatility(self):
        return self._volatility

    @volatility.setter
    def volatility(self, value):
        if (value <= 0):
            raise ValueError("Invalid volatility!")
        self._volatility = value
        self.reset()

    ############################################################################
    @property
    def stock_price(self):
        return self._stock_price

    @stock_price.setter
    def stock_price(self, value):
        if (value < 0):
            raise ValueError("Invalid stock price!")
        self._stock_price = value
        self.reset()

    ############################################################################
    @property
    def interest_rate(self):
        return self._interest_rate

    @interest_rate.setter
    def interest_rate(self, value):
        if (value < 0):
            raise ValueError("Invalid interest rate!")
        self._interest_rate = value
        self.reset()

    ############################################################################
    @property
    def time_to_maturity(self):
        return self._time_to_maturity

    @time_to_maturity.setter
    def time_to_maturity(self, value):
        if (value < 0):
            raise ValueError("Invalid time to maturity!")
        self._time_to_maturity = value
        self.reset()

    ############################################################################
    # Outputs                                                                  #
    ############################################################################
    @property
    def d1(self):
        if (self._d1 is None): self.calculate_d1()
        return self._d1

    def calculate_d1(self):
        S = self.stock_price
        K = self.strike
        T = self.time_to_maturity
        r = self.interest_rate
        q = self.dividend
        sigma = self.volatility
        if (S == 0 or T == 0):
            self._d1 = float("NaN")
        else:
            self._d1 = (math.log(S/K) + (r - q + (sigma**2)/2)*T) \
                     / (sigma*math.sqrt(T))

    ############################################################################
    @property
    def d2(self):
        if (self._d2 is None): self.calculate_d2()
        return self._d2

    def calculate_d2(self):
        T = self.time_to_maturity
        sigma = self.volatility
        self._d2 = self.d1 - sigma*math.sqrt(T)

    ############################################################################
    @property
    def delta(self):
        if (self._delta is None): self.calculate_delta()
        return self._delta

    def calculate_delta(self):
        S = self.stock_price
        T = self.time_to_maturity
        q = self.dividend
        if (S == 0 or T == 0):
            self._delta = float("NaN")
        else:
            if (self.type == "CALL"):
                self._delta = math.exp(-q*T)*norm.cdf(self.d1)
            else: # PUT
                self._delta = -math.exp(-q*T)*norm.cdf(-self.d1)

    ############################################################################
    @property
    def gamma(self):
        if (self._gamma is None): self.calculate_gamma()
        return self._gamma

    def calculate_gamma(self):
        S = self.stock_price
        T = self.time_to_maturity
        q = self.dividend
        sigma = self.volatility
        if (S == 0 or T == 0):
            self._gamma = float("NaN")
        else:
            self._gamma = math.exp(-q*T)*norm.pdf(self.d1) \
                        / (sigma*S*math.sqrt(T))

    ############################################################################
    @property
    def vega(self):
        if (self._vega is None): self.calculate_vega()
        return self._vega

    def calculate_vega(self):
        S = self.stock_price
        T = self.time_to_maturity
        q = self.dividend
        self._vega = math.exp(-q*T)*S*math.sqrt(T)*norm.pdf(self.d1)

    ############################################################################
    @property
    def theta(self):
        if (self._theta is None): self.calculate_theta()
        return self._theta

    def calculate_theta(self):
        S = self.stock_price
        K = self.strike
        T = self.time_to_maturity
        r = self.interest_rate
        q = self.dividend
        sigma = self.volatility
        if (S == 0 or T == 0):
            self._theta = float("NaN")
        else:
            if (self.type == "CALL"):
                self._theta = -math.exp(-q*T)*S*norm.pdf(self.d1)*sigma \
                            / (2*math.sqrt(T)) \
                            + q*math.exp(-q*T)*S*norm.cdf(self.d1) \
                            - r*K*math.exp(-r*T)*norm.cdf(self.d2)
            else: # PUT
                self._theta = -math.exp(-q*T)*S*norm.pdf(self.d1)*sigma \
                            / (2*math.sqrt(T)) \
                            - q*math.exp(-q*T)*S*norm.cdf(-self.d1) \
                            + r*K*math.exp(-r*T)*norm.cdf(-self.d2)

    ############################################################################
    @property
    def price(self):
        if (self._price is None): self.calculate_price()
        return self._price

    def calculate_price(self):
        S = self.stock_price
        K = self.strike
        T = self.time_to_maturity
        r = self.interest_rate
        q = self.dividend
        if (self.type == "CALL"):
            if (S == 0):
                self._price = 0
            else:
                if (T == 0):
                    self._price = max(S - K, 0)
                else:
                    self._price = math.exp(-q*T)*S*norm.cdf(self.d1) \
                                - math.exp(-r*T)*K*norm.cdf(self.d2)
        else: # PUT
            if (S == 0):
                self._price = K*math.exp(-r*T)
            else:
                if (T == 0):
                    self._price = max(K - S, 0)
                else:
                    self._price = math.exp(-r*T)*K*norm.cdf(-self.d2) \
                                - math.exp(-q*T)*S*norm.cdf(-self.d1)

    ############################################################################
    @property
    def parity(self):
        if (self._parity is None): self.calculate_parity()
        return self._parity

    def calculate_parity(self):
        parity = copy.copy(self)
        parity._type = "PUT" if (self.type == "CALL") else "CALL"
        parity._price = None
        parity._delta = None
        parity._theta = None
        parity._parity = self
        self._parity = parity


@xw.func
def opção_delta(cod):
    data = FuncoesDatas()
    hoje = date.today()
    d1 = data.workday(data = data.hoje(),n_dias = -1,feriados_brasil=True, feriados_eua=False)
    #ativo_bds = bds.serie_cadastro('22602225')
    ativos_info = bawm.dados_opcoes(cod)
    cod_bds = ativos_info['BloombergTk'].values[0]
    cod_bds = cod_bds.replace('BDS:','')
    cod_bds = int(cod_bds)
    preco_vol = bds.preco_vol_opt_rv(cod_bds,d1)
    ativo_objeto = ativos_info['FutObj'].values[0]
    futuros = bawm.dados_futuro(ativo_objeto,d1)
    CDI = bawm.in_historico(index_cod ='CDITX' , data_ini=d1, data_fim=hoje)
    cdi_atual = CDI['Valor'].values[0]
    
    tipo = ativos_info['TipoOpt'].values[0]
    tipo = tipo.strip()
    strike = ativos_info['Strike'].values[0]
    dividend = 0
    volatility = preco_vol['vol'].values[0]
    stock_price = futuros['Preco'].values[0]
    time_to_maturity = futuros['DU'].values[0]
    interest_rate = cdi_atual/100
    
    option = Option(
        type = tipo,
        strike = strike,
        dividend = dividend,
        volatility = volatility,
        stock_price = stock_price,
        interest_rate = interest_rate,
        time_to_maturity = (time_to_maturity/252)
    )
    
    return option.delta

def obter_carteira_simulada(df_posicao_dm1:pd.DataFrame, cad_produtos:pd.DataFrame, ordens=pd.DataFrame(), id_tipo_portfolio=None, pos_trade:bool=False,
                            guid_produto_liquidacao:str=None, manter_colunas_adicionais:list=[], base_dm1=None, base_crm=None) -> pd.DataFrame:
    """
    Função criada para o pré-trade, mas adaptada para outros usos

    Parameters
    ----------
    df_posicao_dm1 : pd.DataFrame
        Posição do portfolio.
    cad_produtos : pd.DataFrame
        Dataframe com o cadastro dos produtos.
    ordens : TYPE, optional
        Dataframe com as ordens. The default is pd.DataFrame().
    id_tipo_portfolio : TYPE, optional
        DESCRIPTION. The default is None.
    pos_trade : bool, optional
        Flag indicando se é calculo do pos-trade. The default is False.
    manter_colunas_adicionais: list, optional
        manter mais colunas do dataframe de posição
    base_dm1 : TYPE, optional
        objeto databaseses PosicaoDm1 / PosicaoDm1Pickle. The default is None.
    base_crm : TYPE, optional
        objeto databaseses Crm. The default is None.

    Returns
    -------
    posicao_adj : pd.DataFrame
        DESCRIPTION.

    """
    from objetos import Ativo
    if base_dm1:
        dm1 = base_dm1
    else:
        dm1 = PosicaoDm1Pickle()
    if base_crm:
        crm = base_crm
    else:
        crm = Crm()
    
    colunas_exp = ['Original', 'FinanceiroFinal','FinanceiroFuturo','Movimentacao']
    colunas_ini = ['GuidProduto', 'QuantidadeFinal', 'FinanceiroFinal','FinanceiroFuturo', 'QtdeBloq']         # , 'FinBloq'
            
    
    # Inicializar posição ajustada
    posicao_adj = df_posicao_dm1.copy()
    posicao_adj['QtdeBloq'] = posicao_adj['QtdeBloq'].fillna(0)
    # posicao_adj['FinBloq'] = posicao_adj.apply(lambda x: x['QtdeBloq'] * x['PrecoInicial'], axis=1)
    
    # Tratamento do Caixa
    # -------------------------
    ativo_brl = 'c6999643-98b7-df11-85ec-d8d385b9752e'
    # ativo esquisito que vem como caixa
    ativo_brl_2 = 'ddd05246-c7c1-e511-9f22-005056912b96' # So remover
    posicao_adj = posicao_adj[~posicao_adj['GuidProduto'].isin([ativo_brl_2])]
    
    # Ajustar caixa futuro para ser igual ao atual
    if ativo_brl in posicao_adj['GuidProduto'].unique():
        temp = posicao_adj[posicao_adj['GuidProduto'].isin([ativo_brl])].index
        posicao_adj.loc[temp, 'FinanceiroFuturo'] = posicao_adj.loc[temp, 'FinanceiroFinal']    
    ativo_caixa = ativo_brl
    if id_tipo_portfolio == 1:
        if guid_produto_liquidacao:
            # Se houver um fundo de zeragem cadastrado, transfere a posição de caixa para ele
            ativo_caixa = guid_produto_liquidacao
            temp = posicao_adj[posicao_adj['GuidProduto'].isin([ativo_brl])].index
            posicao_adj.loc[temp, 'GuidProduto'] = ativo_caixa
        if 'COMP' in list(posicao_adj['TipoProduto'].unique()):
            # Soma compromissadas ao caixa
            temp = posicao_adj[posicao_adj['TipoProduto'].isin(['COMP'])].index
            posicao_adj.loc[temp, 'GuidProduto'] = ativo_caixa
            
    posicao_adj = posicao_adj[colunas_ini].groupby('GuidProduto').sum()
    posicao_adj['Movimentacao'] = 0
    posicao_adj.insert(len(posicao_adj.columns), 'Original', posicao_adj['FinanceiroFinal'])        
        
    
    # Ajustar iterativamente as posições financeiras
    net_caixa = 0
    if not ordens.empty:
        for i,ordem in ordens.iterrows():
            ativo = ordem['AtivoGuid']
            
            if ativo not in posicao_adj.index:
                posicao_adj.loc[ativo] = None
                posicao_adj.loc[ativo,'Original'] = 0
                posicao_adj.loc[ativo,'FinanceiroFinal'] = 0
                posicao_adj.loc[ativo,'FinanceiroFuturo'] = 0
                posicao_adj.loc[ativo,'Movimentacao'] = 0
                posicao_adj.loc[ativo,'QuantidadeFinal'] = 0
            
            if ordem['Quantidade'] == 0 and (ordem['Financeiro'] == 0 or not ordem['Financeiro']):
                # Indica que boleta foi "apagada" - provavelmente pelo dpm
                pass
            
            elif ordem['QouF'] == 'Q' and (ordem['Financeiro'] != 0 or ordem['ValorAdj'] != 0):
                pu_ordem = abs(ordem['ValorAdj']) / ordem['Quantidade']
                pu_pos = abs(posicao_adj.loc[ativo,'FinanceiroFinal']) / posicao_adj.loc[ativo,'QuantidadeFinal']
                if pu_ordem != pu_pos:
                    posicao_adj.loc[ativo,'FinanceiroFinal'] = round(posicao_adj.loc[ativo,'QuantidadeFinal'] * pu_ordem,2)
                    posicao_adj.loc[ativo,'FinanceiroFuturo'] = round(posicao_adj.loc[ativo,'QuantidadeFinal'] * pu_ordem,2)
            elif ordem['QouF'] == 'Q' and ordem['ValorAdj'] == 0:
                sinal = 1
                if ordem['ValorAdj']  < 0:
                    sinal = -1
                pu_pos = abs(posicao_adj.loc[ativo,'FinanceiroFinal']) / posicao_adj.loc[ativo,'QuantidadeFinal']                
                # if abs(ordem['ValorAdj']) < 0.9 * pu_pos * ordem['Quantidade']:
                if not math.isnan(pu_pos):
                    ordem['ValorAdj'] = ordem['Quantidade'] * pu_pos                
                    ordem['ValorAdj'] = ordem['ValorAdj'] * sinal                    
            
            posicao_adj.loc[ativo,'Movimentacao'] = ordem['ValorAdj']

            if ordem['ResgTot']:
                posicao_adj.loc[ativo,'FinanceiroFinal'] = 0
                posicao_adj.loc[ativo,'FinanceiroFuturo'] = 0
            elif ordem['QouF'] == 'F':
                # Arredonda o valor para não zerar em resgates totais mal boletados
                posicao_adj.loc[ativo,'FinanceiroFinal'] += round(ordem['ValorAdj'], 0)
                posicao_adj.loc[ativo,'FinanceiroFuturo'] += round(ordem['ValorAdj'], 0)
            else:
                posicao_adj.loc[ativo,'FinanceiroFinal'] += round(ordem['ValorAdj'],2)
                posicao_adj.loc[ativo,'FinanceiroFuturo'] += round(ordem['ValorAdj'],2)
        
            teste_caixa = True
            if 'TipoProduto' in ordens.columns:
                if ordem['TipoProduto'] in ['FUT', 'SWAP']:
                    teste_caixa = False
            if 'IdSysDestino' in ordens.columns:
                # Não conta como caixa boletas que vão pro secundário interno
                if ordem['IdSysDestino'] in [19, 69, 81]:
                    teste_caixa = False
            
            if teste_caixa:
                net_caixa -= ordem['ValorAdj']
            
    if id_tipo_portfolio == 1:            
        if not ativo_caixa in posicao_adj.index:        
            # inserir uma linha de caixa: series + concat
            posicao_adj.loc[ativo_caixa] = None            
            posicao_adj.loc[ativo_caixa,'Original'] = 0
        
        posicao_adj.loc[ativo_caixa,'FinanceiroFinal'] += net_caixa
        posicao_adj.loc[ativo_caixa,'FinanceiroFuturo'] += net_caixa
        posicao_adj.loc[ativo_caixa,'Movimentacao'] = net_caixa
    else:
        if ativo_caixa in posicao_adj.index:
            posicao_adj.drop(ativo_caixa, inplace=True)
    
    # Selecionar colunas finais
    posicao_adj = posicao_adj[colunas_exp]
    
    # Troca ativos do secundario interno por guid produto, se disponível
    sub_pos = posicao_adj[(posicao_adj.index.str.contains(r"sub:sec")) | (posicao_adj.index.str.contains(r"sub:SEC"))]
    for idx, row in sub_pos.iterrows():
        at = Ativo(idx, base_dm1=dm1, base_crm=crm)
        if at.sec_guidproduto:
            posicao_adj.rename(index={idx:at.sec_guidproduto},inplace=True)
        # pos.loc[idx, 'GuidProduto'] = at.sec_guidproduto
    
    # tira ajustes                
    if 'f578ce29-69de-dd11-b24f-000f1f6a9c1c' in posicao_adj.index:
        posicao_adj.drop('f578ce29-69de-dd11-b24f-000f1f6a9c1c', inplace=True)
    
    # Fillna
    for col in colunas_exp:
        posicao_adj[col] = posicao_adj[col].fillna(0)
    
    # Popular tabela de posição com as infos necessárias para verificação pré-trade
    posicao_adj = pd.merge(left=posicao_adj, left_on='GuidProduto', how='left', right=cad_produtos, right_index=True)
    if ativo_caixa in posicao_adj.index:
        posicao_adj.loc[ativo_caixa, 'TipoProduto'] = 'COMP'
        posicao_adj.loc[ativo_caixa, 'GuidEmissor'] = '3ca7c6d6-c603-e011-82b1-d8d385b9752e' # '3de2eeb2-565c-de11-90dd-0003ffe6d283' # Bradesco
    
    # posicao_adj.reset_index(drop=True, inplace=True)
    df = posicao_adj[posicao_adj.index.str.contains("sub:")]
    if len(df) > 0:            
        for idx, row in df.iterrows():
            at = Ativo(id_ativo = idx, base_dm1=dm1, base_crm=crm)
            posicao_adj.loc[idx, 'TipoProduto'] = at.tipo_produto
            posicao_adj.loc[idx, 'Classe'] = at.classe
            posicao_adj.loc[idx, 'NomeProduto'] = at.nome
            posicao_adj.loc[idx, 'SubClasse'] = at.sub_classe
            posicao_adj.loc[idx, 'RiskScore'] = at.risk_score
            posicao_adj.loc[idx, 'GuidEmissor'] = at.emissor_guid                
            posicao_adj.loc[idx, 'Coupon'] = at.coupon
            posicao_adj.loc[idx, 'Taxa'] = at.percentual
    
    df = posicao_adj[posicao_adj['NomeProduto'].isnull()]
    if not df.empty:
        df = df.loc[df.index!='00000000-0000-0000-0000-000000000000']
        for idx, row in df.iterrows():
            at = Ativo(id_ativo = idx, base_dm1=dm1, base_crm=crm)
            posicao_adj.loc[idx, 'TipoProduto'] = at.tipo_produto
            posicao_adj.loc[idx, 'Classe'] = at.classe
            posicao_adj.loc[idx, 'NomeProduto'] = at.nome
            posicao_adj.loc[idx, 'SubClasse'] = at.sub_classe
            posicao_adj.loc[idx, 'RiskScore'] = at.risk_score
            posicao_adj.loc[idx, 'GuidEmissor'] = at.emissor_guid
    
    if pos_trade:
        posicao_adj['Original'] = 0
    posicao_adj['DicioCad'] = posicao_adj['DicioCad'].fillna('NA')
    
    
    return posicao_adj

if __name__ == '__main__':         
    delta = opção_delta('IBOVR92_2023')
    print(delta)    
        
    
        
